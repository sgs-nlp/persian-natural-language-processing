import datetime
import json
import os
import openpyxl
from redis.connection import Encoder
from background_task import background
import datetime

from django.db import models

from conf.settings import BASE_DICT

from nvd.uploader import get_file_upload_to
from nvd.hasher import list_hash, string_hash
from nvd.pre_processing import tokenizer, normilizer
from nvd.storage import data2zip

from conf.settings import DB2FILE_ROOT


class Corpus(models.Model):
    name = models.CharField(
        max_length=128,
        blank=False,
        null=False,
        unique=True,
    )
    file_path = models.CharField(
        max_length=1024,
        null=False,
        blank=False,
    )
    documents = models.ManyToManyField(
        to='Document',
    )
    hash_code = models.CharField(
        max_length=512,
        unique=True,
        null=False,
        blank=False
    )
    is_completed = models.BooleanField(
        default=False,
    )

    def __str__(self):
        return self.name

    def __iter__(self):
        yield 'id', self.pk
        yield 'name', self.name
        yield 'documents', [
            dict(doc).update({'tag': [tag.title for tag in doc.tag_set.all()]}) for doc in self.documents.all()
        ]

    _json_file_path = models.FilePathField(
        null=True,
        blank=True,
    )

    @property
    def json_file_path(self):
        if self._json_file_path is not None:
            if not os.path.isfile(self._json_file_path):
                return self._json_file_path
        path = os.path.join(DB2FILE_ROOT, 'prerequisite', 'corpus')
        zfile_name = f'{self.name}.nvd.zip'
        self._json_file_path = data2zip(dict(self), path, zfile_name)
        return self._json_file_path

    def _words_list(self, string: str) -> list:
        string = normilizer(string)
        sents = tokenizer(string)
        res = []
        for words in sents:
            for word in words:
                word = Word(string=word)
                word.save()
                res.append(word.pk)
        return res

    def _documents_list(self, corpus: list) -> list:
        res = []
        for doc in corpus:
            title = self._words_list(doc.title)
            content = self._words_list(doc.content)
            document = Document()
            document.title__ = title
            document.content__ = content
            document.save()
            tag = Tag()
            tag.document = document
            tag.title = doc.tag
            tag.save()
            res.append(document.pk)
        return res

    def _load_data_from_file(self, file_path):
        wb_obj = openpyxl.load_workbook(file_path)
        sheet = wb_obj.active
        headers = [header.value.lower() for header in list(sheet.iter_rows())[0]]
        corpus = []
        for row in list(sheet.iter_rows())[1:]:
            t = self._TextStruct()
            for header, cell in zip(headers, row):
                if header == 'title':
                    t.title = cell.value
                elif header == 'content':
                    t.content = cell.value
                elif header == 'tag':
                    t.tag = cell.value

            corpus.append(t)
        return corpus

    class _TextStruct:
        def __init__(self, title: str = None, content: str = None, tag: str = None):
            self.title = title
            self.content = content
            self.tag = tag

        def __str__(self):
            return f'title:{self.title}\ncontent: {self.content}\n category: {self.tag}'

    def _set_params(self, other):
        self.pk = other.pk
        # self.name = other.name
        # self.file_path = other.file_path
        # [self.documents.add(document) for document in other.documents.all()]
        # self.hash_code = other.hash_code
        # self.is_completed = other.is_completed

    @property
    def _is_exist(self) -> bool:
        self.name = os.path.basename(self.file_path)
        self.hash_code = string_hash(self.name)
        corpus = BASE_DICT.get_item(self.hash_code)
        if corpus:
            self._set_params(corpus)
            return True
        corpus = Corpus.objects.filter(hash_code=self.hash_code).first()
        if corpus:
            if corpus.is_completed:
                self._set_params(corpus)
                BASE_DICT.set_item(corpus.hash_code, corpus)
                return True
            else:
                corpus.delete()
                return False
        return False

    def save(self, *args, **kwargs):
        if not self._is_exist:
            super(Corpus, self).save(*args, **kwargs)

            create_documents(self.pk)
            # corpus = self._load_data_from_file(self.file_path)
            # documents = self._documents_list(corpus)
            # self.documents.add(*documents)
            # Corpus.objects.filter(pk=self.pk).update(is_completed=True)
            # BASE_DICT.set_item(self.hash_code, self)
            # tmp = self.json_file_path

    def update(self, *args, **kwargs):
        self._json_file_path = None
        tmp = self.json_file_path
        super().update(*args, **kwargs)


class Word(models.Model):
    string = models.CharField(
        max_length=64,
        blank=False,
        null=False,
        unique=True,
    )
    hash_code = models.CharField(
        max_length=512,
        unique=True,
        null=False,
        blank=False
    )

    def __iter__(self):
        yield 'id', self.pk
        yield 'string', self.string

    def _set_params(self, other):
        self.pk = other.pk
        self.string = other.string
        self.hash_code = other.hash_code

    @property
    def _is_exist(self) -> bool:
        self.hash_code = string_hash(self.string)
        word = BASE_DICT.get_item(self.hash_code)
        if word:
            self._set_params(word)
            return True
        word = Word.objects.filter(hash_code=self.hash_code).first()
        if word:
            self._set_params(word)
            BASE_DICT.set_item(word.hash_code, word)
            return True
        return False

    def save(self, *args, **kwargs):
        if not self._is_exist:
            super(Word, self).save(*args, **kwargs)
            BASE_DICT.set_item(self.hash_code, self)


class Document(models.Model):
    title__: list = None
    title = models.ManyToManyField(
        to='Word',
        related_name='words_document4title',
    )
    content__: list = None
    content = models.ManyToManyField(
        to='Word',
        related_name='words_document4abstract',
    )
    hash_code = models.CharField(
        max_length=512,
        unique=True,
        null=False,
        blank=False
    )
    is_completed = models.BooleanField(
        default=False,
    )

    def __iter__(self):
        yield 'id', self.pk
        yield 'title', [dict(word) for word in self.title.all()]
        yield 'content', [dict(word) for word in self.content.all()]

    def _set_params(self, other):
        self.pk = other.pk
        # [self.title.add(word) for word in other.title.all()]
        # [self.content.add(word) for word in other.content.all()]
        # self.hash_code = other.hash_code
        # self.is_completed = other.is_completed

    @property
    def _is_exist(self):
        self.hash_code = list_hash(self.title__ + self.content__)
        document = BASE_DICT.get_item(self.hash_code)
        if document:
            self._set_params(document)
            return True
        document = Document.objects.filter(hash_code=self.hash_code).first()
        if document:
            if document.is_completed:
                self._set_params(document)
                BASE_DICT.set_item(document.hash_code, document)
                return True
            else:
                document.delete()
                return False
        return False

    def save(self, *args, **kwargs):
        if not self._is_exist:
            super(Document, self).save(*args, **kwargs)
            [self.title.add(word) for word in self.title__]
            [self.content.add(word) for word in self.content__]
            Document.objects.filter(pk=self.pk).update(is_completed=True)
            BASE_DICT.set_item(self.hash_code, self)


class Tag(models.Model):
    document = models.ForeignKey(
        to='Document',
        on_delete=models.CASCADE,
    )
    title = models.CharField(
        max_length=128,
    )


# @background()
def create_documents(corpus_id: int) -> None:
    corpus = Corpus.objects.filter(pk=corpus_id).first()
    if corpus is None:
        return
    corpus_data = _load_data_from_file(corpus.file_path)
    documents = _documents_list(corpus_data)
    corpus.documents.add(*documents)
    corpus.is_completed = True
    tmp = corpus.json_file_path
    corpus.save()
    BASE_DICT.set_item(corpus.hash_code, corpus)


def _load_data_from_file(file_path):
    wb_obj = openpyxl.load_workbook(file_path)
    sheet = wb_obj.active
    headers = [header.value.lower() for header in list(sheet.iter_rows())[0]]
    corpus = []

    for row in list(sheet.iter_rows())[1:]:
        t = _TextStruct()
        for header, cell in zip(headers, row):
            if header == 'title':
                t.title = cell.value
            elif header == 'content':
                t.content = cell.value
            elif header == 'tag':
                t.tag = cell.value

        corpus.append(t)
    return corpus


class _TextStruct:
    def __init__(self, title: str = None, content: str = None, tag: str = None):
        self.title = title
        self.content = content
        self.tag = tag

    def __str__(self):
        return f'title:{self.title}\ncontent: {self.content}\n category: {self.tag}'


def _documents_list(corpus: list) -> list:
    res = []
    for doc in corpus:
        title = _words_list(doc.title)
        content = _words_list(doc.content)
        document = Document()
        document.title__ = title
        document.content__ = content
        document.save()
        tag = Tag()
        tag.document = document
        tag.title = doc.tag

        tag.save()
        res.append(document.pk)
    return res


def _words_list(string: str) -> list:
    string = normilizer(string)
    sents = tokenizer(string)
    res = []
    for words in sents:
        for word in words:
            word = Word(string=word)
            word.save()
            res.append(word.pk)
    return res

# def _set_params(self, other):
#     self.pk = other.pk
#     # self.name = other.name
#     # self.file_path = other.file_path
#     # [self.documents.add(document) for document in other.documents.all()]
#     # self.hash_code = other.hash_code
#     # self.is_completed = other.is_completed
#
# @property
# def _is_exist(self) -> bool:
#     self.name = os.path.basename(self.file_path)
#     self.hash_code = string_hash(self.name)
#     corpus = BASE_DICT.get_item(self.hash_code)
#     if corpus:
#         self._set_params(corpus)
#         return True
#     corpus = Corpus.objects.filter(hash_code=self.hash_code).first()
#     if corpus:
#         if corpus.is_completed:
#             self._set_params(corpus)
#             BASE_DICT.set_item(corpus.hash_code, corpus)
#             return True
#         else:
#             corpus.delete()
#             return False
#     return False
